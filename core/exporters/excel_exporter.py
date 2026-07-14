# -*- coding: utf-8 -*-
"""
Exportation en Excel
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os


class ExcelExporter:
    def __init__(self):
        pass

    def export(self, synoptique_data, output_path):
        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Synoptique'

        header_fill = PatternFill(start_color='1F77B4', end_color='1F77B4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = ['Niveau', 'Élément', 'Type', 'Description', 'Distance (m)', 'Fibres']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 10

        row = 2
        tree = synoptique_data.get('tree', {})

        for root_id, root_node in tree.items():
            row = self._add_node_to_excel(ws, root_node, row, 0, border)

        summary_ws = wb.create_sheet('Synthèse')
        summary_ws.cell(row=1, column=1, value='Paramètre').font = header_font
        summary_ws.cell(row=1, column=1).fill = header_fill
        summary_ws.cell(row=1, column=2, value='Valeur').font = header_font
        summary_ws.cell(row=1, column=2).fill = header_fill

        summary_data = [
            ('Style', synoptique_data.get('style', 'N/A')),
            ('Niveaux', synoptique_data.get('levels', 0)),
            ('Éléments totaux', synoptique_data.get('elements_count', 0)),
            ('Chemins', synoptique_data.get('paths_count', 0)),
            ('Distances visibles', 'Oui' if synoptique_data.get('show_distances') else 'Non'),
            ('Fibres visibles', 'Oui' if synoptique_data.get('show_fiber_count') else 'Non'),
        ]

        for i, (param, value) in enumerate(summary_data, 2):
            summary_ws.cell(row=i, column=1, value=param).border = border
            summary_ws.cell(row=i, column=2, value=str(value)).border = border

        wb.save(output_path)

    def _add_node_to_excel(self, ws, node, row, level, border):
        ws.cell(row=row, column=1, value=level).border = border
        ws.cell(row=row, column=2, value=node['name']).border = border
        ws.cell(row=row, column=3, value=node['type']).border = border
        ws.cell(row=row, column=4, value='').border = border
        ws.cell(row=row, column=5, value='').border = border
        ws.cell(row=row, column=6, value='').border = border

        row += 1

        for child in node.get('children', []):
            row = self._add_node_to_excel(ws, child, row, level + 1, border)

        return row
